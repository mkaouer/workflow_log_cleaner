import pandas as pd
import json
from openpyxl import load_workbook
import subprocess
import os

# Input and output file paths
"""
Expected Input File Format:
- The input file should be an Excel (.xlsx) file.
- It must contain a column named 'Error log' where each row contains a CI/CD pipeline failure log.
- The script will analyze each log and generate a summary of the root cause.
- The output file will have an additional column 'llama_8b' containing the AI-generated diagnosis.
"""



def diagnose_ci_cd_failure(error_log):
    """
    Calls the LLaMA 8B model to diagnose CI/CD pipeline failures from logs.
    """
    prompt = f"""
    You are an expert in diagnosing CI/CD pipeline failures from GitHub Actions logs. Analyze the following failure log and summarize the root cause in one sentence. If possible, suggest a fix.

    Failure Log:
    ---
    {error_log}
    ---
    """

    try:
        result = subprocess.run(
            ["ollama", "run", "llama3:8b"],
            input=prompt,  # Keep as a string
            text=True,  # Ensures input is treated as text
            capture_output=True,
            check=True,
            encoding="utf-8"  # 🛠 Forces UTF-8 output handling
        )
        if result.stdout:
            return result.stdout.strip()
        else:
            return "No output from the model."
    except subprocess.CalledProcessError as e:
        return f"Command failed with return code {e.returncode}: {e.stderr}"
    except FileNotFoundError:
        return "Ollama is not installed or not found in PATH."



def process_excel(input_excel_path, output_excel_path):
    """
    Reads the Excel file, processes each row by diagnosing CI/CD failures,
    and appends results in a new column 'llama_8b'. Results are saved after each row.
    """
    df = pd.read_excel(input_excel_path, engine="openpyxl")  # Ensure UTF-8 support

    if 'Error log' not in df.columns:
        print("Error: 'Error log' column not found in input file.")
        return

    # Load existing output file if it exists, otherwise use input file
    try:
        workbook = load_workbook(output_excel_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = load_workbook(input_excel_path)
        sheet = workbook.active

    # Get column index for new column (llama_8b)
    llama_col_index = len(df.columns) + 1

    for index, row in df.iterrows():
        error_log = str(row['Error log']).strip()

        if not error_log:
            result = "No error log available."
        else:
            print(f"Processing row {index + 1}/{len(df)}...")
            try:
                error_log = error_log.encode("utf-8", "ignore").decode("utf-8")  # Remove problematic characters
                result = diagnose_ci_cd_failure(error_log)
            except UnicodeEncodeError as e:
                print(f"Encoding Error in row {index + 1}: {e}")
                result = "Encoding error in log."

        # Save result in the corresponding row of the output file
        sheet.cell(row=index + 2, column=llama_col_index, value=result)

        # Save after each row
        workbook.save(output_excel_path)
        print(f"Row {index + 1} saved to {output_excel_path}")

    print("All rows processed and saved.")


def main():
    input_path = input("Enter the path to the input Excel file: ")
    output_path = input("Enter the path to save the output Excel file: ")
    process_excel(input_path, output_path)


if __name__ == "__main__":
    print("Starting script...")
    main()
    print("Script execution completed.")



